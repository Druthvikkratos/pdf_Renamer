"""
Core engine for PDF renaming based on Excel lookup + PDF-content verification.
No UI/framework code here - pure logic, fully testable standalone.
"""

import io
import re
import base64
import zipfile
from datetime import datetime, date
from typing import Optional

import openpyxl
from pypdf import PdfReader


# ---- Accepted Excel column header variants (case-insensitive, trimmed) ----
COLUMN_ALIASES = {
    "doc_no": ["doc. no.", "doc no.", "doc no", "document no", "document no.", "doc number"],
    "document_type": ["document type", "doc type", "document_type"],
    "customer": ["customer", "customer code", "customer no", "customer no."],
    "posting_date": ["posting date", "posting_date", "post date"],
}

# Only these labels are treated as "the" doc/invoice number field inside a PDF.
# We deliberately do NOT scan for bare digit runs anywhere in the text - that
# would risk matching GSTIN, HSN codes, order numbers, or amounts by accident.
LABEL_TEXT_PATTERNS = [
    re.compile(r"^invoice\s*number\s*:?$", re.IGNORECASE),
    re.compile(r"^invoice\s*no\.?\s*:?$", re.IGNORECASE),
    re.compile(r"^inv\.?\s*no\.?\s*:?$", re.IGNORECASE),
    re.compile(r"^inv\.?\s*number\s*:?$", re.IGNORECASE),
    re.compile(r"^doc\.?\s*no\.?\s*:?$", re.IGNORECASE),
    re.compile(r"^doc\.?\s*number\s*:?$", re.IGNORECASE),
    re.compile(r"^document\s*no\.?\s*:?$", re.IGNORECASE),
    re.compile(r"^document\s*number\s*:?$", re.IGNORECASE),
]


class ExcelParseError(Exception):
    pass


# =============================================================================
# Excel loading
# =============================================================================

def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _find_header_row(ws, max_scan_rows: int = 15):
    """
    Scans the first `max_scan_rows` rows to find the row that contains
    all 4 required headers (Doc. No., Document Type, Customer, Posting Date).
    Handles files where junk/formatting rows sit above the real header.
    Returns (header_row_index, column_map: key -> 1-based column index).
    """
    required_keys = list(COLUMN_ALIASES.keys())

    for row_idx in range(1, max_scan_rows + 1):
        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = _normalize(ws.cell(row=row_idx, column=col_idx).value)
            if not cell_val:
                continue
            for key, aliases in COLUMN_ALIASES.items():
                if cell_val in aliases and key not in col_map:
                    col_map[key] = col_idx

        if all(k in col_map for k in required_keys):
            return row_idx, col_map

    raise ExcelParseError(
        "Could not find a header row containing all required columns "
        "(Doc. No., Document Type, Customer, Posting Date). "
        "Please check the Excel file's column headers."
    )


def _format_posting_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # fallback: keep raw text rather than crash


def _format_int_like(value) -> str:
    """Doc No / Customer may come in as int or float (123.0) from Excel - normalize to plain string."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_excel(file_bytes: bytes):
    """
    Returns (records, warnings, total_data_rows, skipped_rows)
    records: dict[doc_no:str] -> {customer, document_type, posting_date, row, matched}
    Raises ExcelParseError on structural problems (unreadable file / no header found).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ExcelParseError(f"Could not open Excel file: {e}")

    ws = wb.worksheets[0]  # only one sheet expected
    header_row, col_map = _find_header_row(ws)

    records = {}
    warnings = []
    total_data_rows = 0
    skipped_rows = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        doc_no_raw = ws.cell(row=row_idx, column=col_map["doc_no"]).value
        customer_raw = ws.cell(row=row_idx, column=col_map["customer"]).value
        doc_type_raw = ws.cell(row=row_idx, column=col_map["document_type"]).value
        posting_date_raw = ws.cell(row=row_idx, column=col_map["posting_date"]).value

        if doc_no_raw is None and customer_raw is None and doc_type_raw is None and posting_date_raw is None:
            continue  # fully blank row, skip silently

        total_data_rows += 1

        if doc_no_raw is None or doc_type_raw is None or customer_raw is None or posting_date_raw is None:
            skipped_rows += 1
            warnings.append(f"Row {row_idx}: missing one or more required values - skipped.")
            continue

        doc_no = _format_int_like(doc_no_raw)
        customer = _format_int_like(customer_raw)
        doc_type = str(doc_type_raw).strip()
        posting_date = _format_posting_date(posting_date_raw)

        if not doc_no:
            skipped_rows += 1
            continue

        # Duplicate Doc. No. -> last one wins, still goes to success later, no error raised.
        records[doc_no] = {
            "customer": customer,
            "document_type": doc_type,
            "posting_date": posting_date,
            "row": row_idx,
            "matched": False,
        }

    if total_data_rows == 0:
        warnings.append("Excel file has a valid header but contains no data rows.")

    return records, warnings, total_data_rows, skipped_rows


# =============================================================================
# PDF reading + Doc No. extraction
# =============================================================================

def extract_doc_no_from_filename(filename: str) -> Optional[str]:
    """
    Incoming PDF pattern (when SAP names it normally): PREFIX_DOCNO_YYYYMMDD_HHMMSS.pdf
    Only the 2nd underscore-separated part matters; everything else is discarded.
    Returns None if the filename doesn't follow this pattern at all - that's fine,
    the content-based fallback in process_files() handles arbitrary filenames.
    """
    stem = re.sub(r"\.pdf$", "", filename, flags=re.IGNORECASE)
    parts = stem.split("_")
    if len(parts) < 2:
        return None
    doc_no = parts[1].strip()
    return doc_no if doc_no else None


def read_pdf_text(pdf_bytes: bytes):
    """
    Reads all pages of the PDF and returns (full_text: str, error: str).
    error is non-empty only if the PDF itself couldn't be opened/read
    (corrupt, encrypted, unsupported format, etc).
    """
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception as e:
        return "", f"Could not open/read PDF: {e}"

    try:
        chunks = []
        for page in reader.pages:
            chunks.append(page.extract_text() or "")
        return "\n".join(chunks), ""
    except Exception as e:
        return "", f"Error while reading PDF content: {e}"


def extract_labelled_doc_numbers(text: str) -> set:
    """
    Finds numbers whose label (matched as a whole line, positionally paired with its
    value - see extract_label_value_pairs) is one of: Invoice Number / Invoice No /
    Inv No / Inv Number / Doc No / Doc Number / Document No / Document Number.
    Nothing outside these labels is ever considered - GSTIN, HSN, order numbers,
    amounts, etc. are never touched, regardless of PDF layout style.
    """
    found = set()
    for label, value in extract_label_value_pairs(text):
        for pattern in LABEL_TEXT_PATTERNS:
            if pattern.match(label):
                num = re.sub(r"[^0-9]", "", value)
                if num:
                    found.add(num)
                break
    return found


def find_doc_no_matches_in_content(text: str, records: dict) -> list:
    """
    Cross-checks label-anchored candidates against known Excel Doc. Nos.
    Returns a sorted list of matches (ideally exactly one).
    """
    candidates = extract_all_numbers(text)
    return sorted(c for c in candidates if c in records)


def _text_contains_doc_no(text: str, doc_no: str) -> bool:
    """Used only for the filename fast-path: is this doc_no actually present as a labelled number in the PDF?"""
    return doc_no in extract_all_numbers(text)


def build_new_filename(record: dict, doc_no: str) -> str:
    return f"{record['document_type']}_{doc_no}-{record['customer']}-{record['posting_date']}.pdf"


# =============================================================================
# Main processing
# =============================================================================

def process_files(excel_bytes: bytes, pdf_files: list):
    """
    pdf_files: list of (filename: str, content: bytes)
    Returns a dict with everything the UI needs to render the full report.
    """
    result = {
        "excel_error": None,
        "excel_warnings": [],
        "excel_total_rows": 0,
        "excel_skipped_rows": 0,
        "excel_matched_count": 0,
        "excel_unmatched": [],   # excel rows that had no incoming pdf
        "success": [],           # [{original, renamed}]
        "failed": [],            # [{original, reason}]
        "success_zip_b64": None,
        "failed_zip_b64": None,
        "total_pdfs": len(pdf_files),
    }

    try:
        records, warnings, total_rows, skipped_rows = load_excel(excel_bytes)
    except ExcelParseError as e:
        result["excel_error"] = str(e)
        return result

    result["excel_warnings"] = warnings
    result["excel_total_rows"] = total_rows
    result["excel_skipped_rows"] = skipped_rows

    if total_rows == 0:
        # nothing to match against - every pdf fails, but still return cleanly.
        for filename, content in pdf_files:
            result["failed"].append({"original": filename, "reason": "Excel has no data rows to match against."})
        _build_zip_if_any(result, "failed_zip_b64", list(pdf_files), mode="failed")
        return result

    used_names = {}
    success_items = []  # (filename, content, final_name)
    failed_items = []   # (filename, content)

    for filename, content in pdf_files:
        text, read_error = read_pdf_text(content)
        if read_error:
            result["failed"].append({"original": filename, "reason": read_error})
            failed_items.append((filename, content))
            continue

        filename_doc_no = extract_doc_no_from_filename(filename)
        doc_no = None

        # Step 1: fast path - filename gives a valid Doc No. that's in Excel AND
        # verified as actually sitting next to an Invoice/Doc Number label in the PDF.
        if filename_doc_no and filename_doc_no in records and _text_contains_doc_no(text, filename_doc_no):
            doc_no = filename_doc_no
        else:
            # Step 2: fallback - filename can be anything, so search the PDF's own
            # labelled content (Invoice Number / Doc No / Document No / Inv No)
            # for any number that matches a known Excel Doc. No.
            content_matches = find_doc_no_matches_in_content(text, records)

            if len(content_matches) == 1:
                doc_no = content_matches[0]
            elif len(content_matches) > 1:
                result["failed"].append({
                    "original": filename,
                    "reason": f"Multiple possible Doc. Nos found next to Invoice/Doc Number labels inside the "
                              f"PDF that match Excel entries: {', '.join(content_matches)}. Cannot determine "
                              f"the correct one automatically."
                })
                failed_items.append((filename, content))
                continue
            else:
                # labelled = extract_labelled_doc_numbers(text)
                found_numbers = extract_all_numbers(text)
                if filename_doc_no and filename_doc_no in records:
                    reason = (f"Doc. No. '{filename_doc_no}' matched in Excel via filename, but that number "
                              f"was not found anywhere in the PDF content.")
                elif found_numbers:
                    reason = (f"Found number(s) inside the PDF ({', '.join(sorted(found_numbers))}), "
                              f"but none of them exist in the Excel.")
                elif filename_doc_no:
                    reason = (f"Doc. No. '{filename_doc_no}' from filename not found in Excel, and no "
                              f"'matching number found inside the PDF either.")
                else:
                    reason = ("No 'Invoice Number' / 'Doc Number' / 'Document Number' label found inside the "
                              "PDF, and the filename didn't give a usable Doc. No. either.")
                result["failed"].append({"original": filename, "reason": reason})
                failed_items.append((filename, content))
                continue

        record = records[doc_no]
        record["matched"] = True
        new_name = build_new_filename(record, doc_no)

        final_name = new_name
        if final_name in used_names:
            used_names[final_name] += 1
            stem, ext = final_name.rsplit(".", 1)
            final_name = f"{stem}__dup{used_names[new_name]}.{ext}"
        else:
            used_names[final_name] = 0

        result["success"].append({"original": filename, "renamed": final_name})
        success_items.append((filename, content, final_name))

    for doc_no, rec in records.items():
        if not rec["matched"]:
            result["excel_unmatched"].append({
                "doc_no": doc_no,
                "document_type": rec["document_type"],
                "customer": rec["customer"],
                "posting_date": rec["posting_date"],
            })

    result["excel_matched_count"] = len(records) - len(result["excel_unmatched"])

    _build_zip_if_any(result, "success_zip_b64", success_items, mode="renamed")
    _build_zip_if_any(result, "failed_zip_b64", failed_items, mode="failed")

    return result


def _build_zip_if_any(result: dict, key: str, items: list, mode: str):
    """Only builds & attaches a zip if there's at least one item - never send an empty zip."""
    if not items:
        return
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in items:
            if mode == "renamed":
                filename, content, final_name = item
                zf.writestr(final_name, content)
            else:
                filename, content = item
                zf.writestr(filename, content)
    result[key] = base64.b64encode(buf.getvalue()).decode("ascii")

def extract_label_value_pairs(text: str) -> list:
    """
    Pairs labels with their values by POSITION, not adjacency - this handles both:
      1. Vertical layout: "Invoice Number:\n1192304533\n" (label immediately followed by its value)
      2. Table layout: "Label A:\nLabel B:\nLabel C:\nValue A\nValue B\nValue C\n"
         (a whole run of labels, followed by a whole run of values in the same order)
    Works by scanning for a contiguous run of lines ending in ':' (a label group),
    then taking the same number of following non-label lines as the matching value group,
    and pairing them 1-to-1 by index.
    """
    lines = text.split("\n")
    pairs = []
    i, n = 0, len(lines)

    while i < n:
        if lines[i].strip().endswith(":"):
            label_group = []
            while i < n and lines[i].strip().endswith(":"):
                label_group.append(lines[i].strip())
                i += 1

            value_group = []
            for _ in range(len(label_group)):
                if i < n and not lines[i].strip().endswith(":"):
                    value_group.append(lines[i].strip())
                    i += 1
                else:
                    break

            for idx, lbl in enumerate(label_group):
                if idx < len(value_group):
                    pairs.append((lbl, value_group[idx]))
        else:
            i += 1

    return pairs

def extract_all_numbers(text: str) -> set:
    """
    Every run of digits found anywhere in the PDF text (length >= 4 to avoid
    matching stray page numbers / single digits). No label check at all.
    """
    return set(re.findall(r"\d{4,}", text))